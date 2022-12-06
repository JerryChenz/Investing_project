import yfinance
import openpyxl
import shutil
import pathlib
import os
from datetime import datetime
import pandas as pd
import scrap_mod


class Asset:
    """Parent class"""

    def __init__(self, security_code):
        self.security_code = security_code
        self.name = None
        self.price = None
        self.exchange = None
        self.excess_return = None
        self.target_buy = None


class Stock(Asset):
    """child class"""

    def __init__(self, security_code):
        """ """
        super().__init__(security_code)

        self.dividends = None
        self.shares = None
        self.report_currency = None
        self.is_df = None
        self.bs_df = None
        self.next_earnings = None

    def load_from_yf(self):
        """Scrap the data from yahoo finance"""

        ticker_data = yfinance.Ticker(self.security_code)

        self.name = ticker_data.info['shortName']
        self.price = [ticker_data.info['currentPrice'], ticker_data.info['currency']]
        self.dividends = ticker_data.dividends
        self.exchange = ticker_data.info['exchange']
        self.shares = ticker_data.info['sharesOutstanding']
        self.report_currency = ticker_data.info['financialCurrency']
        self.is_df = scrap_mod.get_income_statement(self.security_code)
        self.bs_df = scrap_mod.get_balance_sheet(self.security_code)
        self.next_earnings = pd.to_datetime(datetime.fromtimestamp(ticker_data.info['mostRecentQuarter'])
                                            .strftime("%Y-%m-%d")) + pd.DateOffset(months=6)

    def create_val_xlsx(self):
        """Return a raw_fin_data xlsx for the stock"""

        new_val_name = ""
        new_bool = False

        # Copy the latest Valuation template
        cwd = pathlib.Path.cwd().resolve()
        try:
            template_folder_path = cwd / 'Template'
            if pathlib.Path(template_folder_path).exists():
                template_path_list = [val_file_path for val_file_path in template_folder_path.iterdir()
                                      if template_folder_path.is_dir() and val_file_path.is_file()]
                if len(template_path_list) > 1 or len(template_path_list) == 0:
                    raise FileNotFoundError("The template file error", "temp_file")
            else:
                raise FileNotFoundError("The template folder doesn't exist", "temp_folder")
        except FileNotFoundError as err:
            if err.args[1] == "temp_folder":
                print("The template folder doesn't exist")
            if err.args[1] == "temp_file":
                print("The template file error")
        else:
            new_val_name = self.security_code + " " + os.path.basename(template_path_list[0])
            if not pathlib.Path(new_val_name).exists():
                shutil.copy(template_path_list[0], new_val_name)
                new_bool = True

        # load and update the new valuation xlsx
        wb = openpyxl.load_workbook(new_val_name)
        self.update_dashboard(wb, new_bool)
        self.update_data(wb)

        wb.save(new_val_name)

    def update_dashboard(self, wb, new_bool=False):
        """Update the Dashboard sheet"""

        dash_sheet = wb['Dashboard']
        if new_bool:
            dash_sheet.cell(row=3, column=3).value = self.security_code
            dash_sheet.cell(row=4, column=3).value = self.name
            dash_sheet.cell(row=5, column=3).value = datetime.today().strftime('%Y-%m-%d')
            dash_sheet.cell(row=3, column=8).value = self.exchange
            dash_sheet.cell(row=12, column=8).value = self.report_currency
        dash_sheet.cell(row=6, column=3).value = self.next_earnings
        dash_sheet.cell(row=4, column=8).value = self.price[0]
        dash_sheet.cell(row=4, column=9).value = self.price[1]
        dash_sheet.cell(row=5, column=8).value = self.shares
        dash_sheet.cell(row=13, column=8).value = scrap_mod.get_forex_rate(self.price[1], self.report_currency)

    def update_data(self, wb):
        """Update the Data sheet"""

        data_sheet = wb['Data']
        data_sheet.cell(row=3, column=3).value = self.is_df.columns[0]  # last financial year
        # figures in
        figures_in = int((len(str(self.is_df.iloc[0, 0])) - 9) / 3 + 0.99) * 1000
        data_sheet.cell(row=4, column=3).value = figures_in
        # load income statement
        for i in range(len(self.is_df.columns)):
            data_sheet.cell(row=7, column=i + 3).value = int(self.is_df.iloc[0, i] / figures_in)
            data_sheet.cell(row=9, column=i + 3).value = int(self.is_df.iloc[1, i] / figures_in)
            data_sheet.cell(row=11, column=i + 3).value = int(self.is_df.iloc[2, i] / figures_in)
            data_sheet.cell(row=17, column=i + 3).value = int(self.is_df.iloc[3, i] / figures_in)
            data_sheet.cell(row=18, column=i + 3).value = int(self.is_df.iloc[4, i] / figures_in)
        # load balance sheet
        for i in range(1, len(self.bs_df.columns)):
            data_sheet.cell(row=20, column=i + 3).value = int(self.bs_df.iloc[0, i] / figures_in)
            data_sheet.cell(row=21, column=i + 3).value = int(self.bs_df.iloc[1, i] / figures_in)
            data_sheet.cell(row=22, column=i + 3).value = int(self.bs_df.iloc[2, i] / figures_in)
            data_sheet.cell(row=23, column=i + 3).value = int(self.bs_df.iloc[3, i] / figures_in)
            data_sheet.cell(row=25, column=i + 3).value = int(self.bs_df.iloc[4, i] / figures_in)
            data_sheet.cell(row=26, column=i + 3).value = int(self.bs_df.iloc[5, i] / figures_in)
            data_sheet.cell(row=27, column=i + 3).value = int(self.bs_df.iloc[6, i] / figures_in)
            data_sheet.cell(row=28, column=i + 3).value = int(self.bs_df.iloc[7, i] / figures_in)

    def export_statements(self):
        """Export the income statement and balance sheet"""

        self.is_df.to_csv(f'{self.security_code}_income_statement.csv', sep=',', encoding='utf-8')
        self.bs_df.to_csv(f'{self.security_code}_balance_sheet.csv', sep=',', encoding='utf-8')
