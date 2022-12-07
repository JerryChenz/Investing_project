import openpyxl
import xlwings
import pathlib
import security_mod
import scrap_mod
import yfinance
import pandas as pd
import re


def update_stocks_val(dash_sheet):
    """Update the stock valuations in the pipeline folder"""

    ticker_info = yfinance.Ticker(dash_sheet.range('C3').value).info

    if pd.to_datetime(dash_sheet.range('C5').value) > pd.to_datetime(dash_sheet.range('C6').value):
        dash_sheet.range('E6').value = "Outdated"
    else:
        dash_sheet.range('E6').value = ""
    dash_sheet.range('H4').value = ticker_info['currentPrice']
    # dash_sheet.range('H5').value = ticker_info['sharesOutstanding']
    dash_sheet.range('H13').value = scrap_mod.get_forex_rate(dash_sheet.range('I4').value,
                                                             dash_sheet.range('H12').value)


def instantiate_asset(p):
    """initiate an asset from the valuation file"""

    # get the formula results using xlwings because openpyxl doesn't evaluate formula
    with xlwings.App(visible=False):
        xl_book = xlwings.Book(p)
        dash_sheet = xl_book.sheets('Dashboard')
        # Update the stock_valuations files first in the pipeline folder
        update_stocks_val(dash_sheet)
        # instantiate the assets
        a = security_mod.Asset(dash_sheet.range('C3').value)
        a.name = dash_sheet.range('C4').value
        a.exchange = dash_sheet.range('H3').value
        a.price = dash_sheet.range('H4').value
        a.ideal_price = dash_sheet.range('C22').value
        a.current_irr = dash_sheet.range('H22').value
        a.risk_premium = dash_sheet.range('H23').value
        a.val_status = dash_sheet.range('E6').value
        a.periodic_payment = dash_sheet.range('C19').value
        a.next_earnings = dash_sheet.range('C6').value

    return a


class Pipeline:
    """A pipeline with many assets"""

    def __init__(self):
        self.assets = []

    def load_opportunities(self):
        """Load the asset information from the opportunities folder"""

        # Copy the latest Valuation template
        opportunities_folder_path = pathlib.Path.cwd().resolve() / 'Opportunities'
        r = re.compile(".*Valuation_v")

        try:
            if pathlib.Path(opportunities_folder_path).exists():
                path_list = [val_file_path for val_file_path in opportunities_folder_path.iterdir()
                             if opportunities_folder_path.is_dir() and val_file_path.is_file()]
                opportunities_path_list = list(item for item in path_list if r.match(str(item)))
                if len(opportunities_path_list) == 0:
                    raise FileNotFoundError("No opportunity file", "opp_file")
            else:
                raise FileNotFoundError("The opportunities folder doesn't exist", "opp_folder")
        except FileNotFoundError as err:
            if err.args[1] == "opp_folder":
                print("The opportunities folder doesn't exist")
            if err.args[1] == "opp_file":
                print("No opportunity file", "opp_file")
        else:
            # load and update the new valuation xlsx
            for p in opportunities_path_list:
                # load and update the new valuation xlsx
                self.assets.append(instantiate_asset(p))
            # load the opportunities
            monitor_file_path = opportunities_folder_path / 'Pipeline_monitor' / 'Pipeline_monitor.xlsx'
            monitor_wb = openpyxl.load_workbook(monitor_file_path)
            self.update_monitor(monitor_wb)
            monitor_wb.save(monitor_file_path)

    def update_monitor(self, wb):
        """update the Pipeline_monitor file"""

        # Clear existing data
        monitor_sheet = wb['Monitor']
        for c in monitor_sheet['B5':'I200']:
            for cell in c:
                cell.value = None

        r = 5
        for a in self.assets:
            monitor_sheet.cell(row=r, column=2).value = a.security_code
            monitor_sheet.cell(row=r, column=3).value = a.name
            monitor_sheet.cell(row=r, column=4).value = a.exchange
            monitor_sheet.cell(row=r, column=5).value = a.price
            monitor_sheet.cell(row=r, column=6).value = a.current_irr
            monitor_sheet.cell(row=r, column=7).value = a.risk_premium
            monitor_sheet.cell(row=r, column=8).value = f'=F{r}-G{r}'
            monitor_sheet.cell(row=r, column=9).value = a.periodic_payment
            monitor_sheet.cell(row=r, column=10).value = f'=I{r}/E{r}'
            monitor_sheet.cell(row=r, column=11).value = a.ideal_price
            monitor_sheet.cell(row=r, column=12).value = a.next_earnings
            monitor_sheet.cell(row=r, column=13).value = a.val_status
            r += 1
